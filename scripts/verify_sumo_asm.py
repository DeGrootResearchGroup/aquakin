"""Verify the SUMO-imported ASM models against their source spreadsheets.

Cross-checks each shipped ASM model, term-by-term and value-by-value, against
the SUMO ``Process code/Model base/Museum/*.xlsm`` source it was generated from:

1. **Parameter values** -- every aquakin parameter vs the SUMO Parameters sheet.
2. **Rate-term constants** -- for each process, the half-saturation constant of
   each Monod saturation/inhibition term (catches the per-group constant
   ``collapse`` -- e.g. a nitrifier term inheriting the heterotroph value).
3. **Stoichiometry structure** -- the set of species participating in each
   process (catches dropped species and biomass swaps).
4. **Stoichiometric coefficients** -- every numeric coefficient, evaluated from
   the SUMO symbolic cell with SUMO parameter values, including the
   charge-balance (``SALK``/``SHCO``) cross-reference terms.

The SUMO spreadsheets are not redistributable, so they are not in this repo;
point this tool at a local copy. Requires ``openpyxl`` (not a runtime
dependency).

Usage::

    python scripts/verify_sumo_asm.py "/path/to/SUMO/Process code/Model base/Museum"

Exit status is non-zero if any discrepancy is found.
"""

from __future__ import annotations

import re
import sys

import numpy as np

import aquakin

MODELS = {
    "asm2d": "ASM2D.xlsm",
    "asm2d_tud": "ASM2D_TUD.xlsm",
    "asm3": "ASM3.xlsm",
    "asm3_biop": "ASM3_BioP.xlsm",
}
_NON_SPECIES = {"j", "Symbol", "Name", "Rate", "Unit", "Reaction", "Rule"}


def _pn(s):
    """SUMO symbol -> a comparable token (Greek spelled out)."""
    return str(s).strip().replace("μ", "mu").replace("η", "eta").replace("θ", "theta")


def _norm(s):
    """Aggressive normalisation for matching parameter / process names."""
    return re.sub(r"[^a-z0-9]", "", _pn(s).lower())


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _load_sumo(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    params = {}
    for row in wb["Parameters"].iter_rows():
        nz = [c.value for c in row if c.value is not None]
        if len(nz) >= 3 and isinstance(nz[0], str):
            v = _to_float(nz[2])
            if v is not None:
                params[_pn(nz[0])] = v
    ws = wb["Model"]
    cols = {c.column: c.value for c in ws[3]
            if c.value and c.value not in _NON_SPECIES}
    species = list(cols.values())
    procs = {}
    for r in range(4, ws.max_row + 1):
        nm = ws.cell(r, 4).value
        j = ws.cell(r, 2).value
        if not nm:
            continue
        cells = {sp: str(ws.cell(r, col).value)
                 for col, sp in cols.items()
                 if ws.cell(r, col).value not in (None, "")}
        rate = ws.cell(r, [c.column for c in ws[3] if c.value == "Rate"][0]).value
        if cells:
            procs.setdefault(_norm(nm), (nm, j, cells, str(rate)))
    return params, species, procs


def _eval_cell(expr, params, vrefs):
    e = _pn(expr)
    subst = {**{_pn(k): v for k, v in vrefs.items()}, **params}
    for k in sorted(subst, key=len, reverse=True):
        e = e.replace(k, repr(subst[k]))
    if re.search(r"[A-Za-z]{2,}", e):
        return None
    try:
        return eval(e, {"__builtins__": {}}, {})  # noqa: S307 - arithmetic only
    except Exception:
        return None


def _parse_sumo_monod(rate, species):
    out = []
    for fac in rate.split("*"):
        fac = fac.strip()
        for pre, ft in (("Msat", "sat"), ("Minh", "inh")):
            if fac.startswith(pre) and not fac.startswith("MR"):
                rest = fac[len(pre):]
                sp = next((s for s in sorted(species, key=len, reverse=True)
                           if rest.startswith(s)), None)
                if sp:
                    out.append((ft, sp, _norm(rest[len(sp):].lstrip(","))))
                break
    return out


def _aquakin_monod(rate_expr, exprs):
    texts, seen = [rate_expr], set()
    i = 0
    while i < len(texts):
        for tok in re.findall(r"[A-Za-z_]\w*", texts[i]):
            if tok in exprs and tok not in seen:
                seen.add(tok)
                texts.append(exprs[tok])
        i += 1
    out = []
    for t in texts:
        for m in re.finditer(r"monod(_inh)?\(\s*\[(\w+)\]\s*,\s*(\w+)\s*\)", t):
            out.append(("inh" if m.group(1) else "sat", m.group(2), _norm(m.group(3))))
    return out


def verify(base_dir):
    import importlib.resources as ir

    total_issues = 0
    for name, xlsm in MODELS.items():
        net = aquakin.load_model(name)
        aqpar = {_norm(p): float(net.parameter_values({})[net.param_index[p]])
                 for p in net.param_index}
        S = np.asarray(net.compute_stoich(net.default_parameters()))
        rn = {_norm(n): i for i, n in enumerate(net.reaction_names)}
        si = net.species_index
        txt = (ir.files("aquakin.models") / f"{name}.yaml").read_text()
        exprs = dict(re.findall(r'\n  (\w+):\s*"([^"]+)"', txt.split("reactions:")[0]))
        aqproc = {}
        for b in re.split(r"\n  - name: ", txt)[1:]:
            nm = b.split("\n", 1)[0].strip()
            m = re.search(r'rate:\s*"([^"]+)"', b)
            sto = re.search(r"stoichiometry:\n((?:\s+\w+:.*\n)+)", b)
            keys = set(re.findall(r"^\s+(\w+):", sto.group(1), re.M)) if sto else set()
            aqproc[_norm(nm)] = (nm, m.group(1) if m else "", keys)

        spar, species, sproc = _load_sumo(f"{base_dir}/{xlsm}")
        issues = []

        # 1. parameter values
        for k, av in aqpar.items():
            if k in spar and abs(av - spar[k]) > 1e-9 * max(abs(spar[k]), 1) + 1e-12:
                issues.append(f"param {k}: aquakin={av} vs SUMO={spar[k]}")

        for pkey, (snm, j, cells, srate) in sproc.items():
            if pkey not in rn:
                continue
            ri = rn[pkey]
            # 2. rate-term constants
            if pkey in aqproc:
                sd = {(ft, sp): c for ft, sp, c in _parse_sumo_monod(srate, species)}
                ad = {(ft, sp): c for ft, sp, c in
                      _aquakin_monod(aqproc[pkey][1], exprs)}
                for key in sd & ad.keys():
                    sv, av = spar.get(sd[key]), aqpar.get(ad[key])
                    if sv is not None and av is not None and \
                            abs(sv - av) > 1e-9 * max(abs(sv), 1) + 1e-12:
                        issues.append(f"[{snm}] {key[0]}({key[1]}) constant: "
                                      f"aquakin={av} vs SUMO={sv}")
            # 3. stoichiometry structure
            if pkey in aqproc:
                ss = {sp for sp in cells if str(cells[sp]).strip() not in ("", "0")}
                aq_keys = aqproc[pkey][2]
                for sp in (ss & set(si)) - aq_keys:
                    issues.append(f"[{snm}] species {sp} present in SUMO, missing in aquakin")
                for sp in aq_keys - ss:
                    if sp in si:
                        issues.append(f"[{snm}] species {sp} in aquakin, not in SUMO")
            # 4. coefficient values (two-pass for charge-balance cross-refs)
            vals = {sp: _eval_cell(e, spar, {}) for sp, e in cells.items()}
            vrefs = {f"v{j}_{sp}": v for sp, v in vals.items() if v is not None}
            for sp, expr in cells.items():
                if sp not in si:
                    continue
                sval = _eval_cell(expr, spar, vrefs)
                if sval is None:
                    continue
                aval = float(S[ri, si[sp]])
                if abs(sval - aval) > 1e-5 * max(abs(sval), 1) + 1e-8:
                    issues.append(f"[{snm}] {sp} coefficient: "
                                  f"aquakin={aval:.5f} vs SUMO={sval:.5f}")

        status = "OK" if not issues else f"{len(issues)} ISSUES"
        print(f"## {name}: {status}")
        for msg in issues:
            print(f"    {msg}")
        total_issues += len(issues)
    return total_issues


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(1 if verify(sys.argv[1]) else 0)
